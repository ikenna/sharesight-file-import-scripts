#!/usr/bin/env python2


"""
MIT License

Copyright(c) 2020, Ikenna C. Nwaiwu. ikenna4u@gmail.com

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.

"""



import openpyxl
import copy
import argparse
import csv


class Transaction:
    def __init__(self, trade_id, trade_date, instrument_code, market_code, quantity, price, transaction_type,
                 exchange_rate, brokerage, brokerage_currency, comments):
        self.trade_id = trade_id
        self.trade_date = trade_date
        self.instrument_code = instrument_code
        self.market_code = market_code
        self.quantity = quantity
        self.price = price
        self.transaction_type = transaction_type
        self.exchange_rate = exchange_rate
        self.brokerage = brokerage
        self.brokerage_currency = brokerage_currency
        self.comments = comments

    def set_brokerage(self, brokerage):
        self.brokerage = brokerage

    def to_csv(self):
        output = [self.trade_id, self.trade_date, self.instrument_code, self.market_code, self.quantity, self.price,
                  self.transaction_type,
                  self.exchange_rate, self.brokerage, self.brokerage_currency, self.comments]
        return ','.join(output)


def add_commission(dict_of_commissions, trade_id_transaction_dict):
    all_transactions = []
    for key in trade_id_transaction_dict:
        transaction_line = copy.deepcopy(trade_id_transaction_dict.get(key))
        #Note: Brokerage value here is commission + Exchange Fee + Stamp Duty
        commission = dict_of_commissions.get(key + "-Commission", 0)
        exchange_fee = dict_of_commissions.get(key + "-Exchange Fee", 0)
        stamp_duty = dict_of_commissions.get(key + "-UK Stamp Duty", 0)
        brokerage_value = float(commission) + float(exchange_fee) + float(stamp_duty)
        transaction_line.set_brokerage(str(brokerage_value))
        all_transactions.append(transaction_line)
    all_transactions.sort(key=lambda x: x.trade_id)
    return all_transactions


def create_transaction_for_saxo(row):
    trade_id = str(row[0].value)
    trade_date = str(row[3].value.strftime("%Y-%m-%d"))
    symbol = str(row[11].value)
    instrument_code = symbol.split(":")[0]
    exchange = str(row[12].value)
    if exchange == "New York Stock Exchange":
        market_code = "NYSE"
    elif exchange == "London Stock Exchange":
        market_code = "LSE"
    elif exchange == "Toronto Stock Exchange":
        market_code = "TSE"
    elif exchange == "NASDAQ (Small cap)":
        market_code = "NASDAQ"
    else:
        market_code = exchange
    quantity = str(row[6].value)
    price = str(row[7].value)
    buy_or_sell = str(row[4].value)
    if buy_or_sell == "Bought":
        transaction_type = "BUY"
    else:
        transaction_type = "SELL"
    brokerage = "?"
    brokerage_currency = "GBP"
    comments = ""
    booked_amount = float(row[10].value)
    trade_value = float(row[8].value)
    if booked_amount != 0:
        exchange_rate = str(trade_value / booked_amount)
    else:
        exchange_rate = "0"
    transaction = Transaction(trade_id, trade_date, instrument_code, market_code, quantity, price, transaction_type,
                              exchange_rate, brokerage, brokerage_currency, comments)
    return transaction

def create_transaction_for_ig(row):
    trade_id = str(row[17])
    trade_date = str(row[0])
    instrument_code = str(row[3])
    market_code = ""
    quantity = str(row[5])
    price = str(row[6])
    transaction_type = str(row[4])
    brokerage = str(float(row[10]) + float(row[9]))
    brokerage_currency = "USD"
    comments = ""
    consideration_in_usd = float(row[8])
    charges_in_usd = float(row[10])
    cost_proceeds_in_gbp = float(row[11])
    exchange_rate = str((consideration_in_usd + charges_in_usd) / cost_proceeds_in_gbp)
    transaction = Transaction(trade_id, trade_date, instrument_code, market_code, quantity, price, transaction_type,
                              exchange_rate, brokerage, brokerage_currency, comments)
    return transaction

def get_trade_id_to_commission_dict(work_book):
    trade_booked_amount_sheet = work_book.get_sheet_by_name('Trade Booked Amount')
    dict_of_commissions = {}
    for row in trade_booked_amount_sheet.iter_rows():
        if str(row[8].value) == "Commission":
            dict_of_commissions.update({str(row[0].value) + "-" + str(row[8].value): str(row[10].value * -1)})
        if str(row[8].value) == "Exchange Fee":
            dict_of_commissions.update({str(row[0].value) + "-" + str(row[8].value): str(row[10].value * -1)})
        if str(row[8].value) == "UK Stamp Duty":
            dict_of_commissions.update({str(row[0].value) + "-" + str(row[8].value): str(row[10].value * -1)})
    return dict_of_commissions


def get_trade_id_transaction_dict(trades_sheet):
    trade_id_transaction_dict = {}
    for row in trades_sheet.iter_rows():
        if row[0].value != "Trade ID":
            transaction = create_transaction_for_saxo(row)
            trade_id_transaction_dict.update({transaction.trade_id: transaction})
    return trade_id_transaction_dict


def main_saxo(excel_file):
    work_book = openpyxl.load_workbook(excel_file)
    trades_sheet = work_book.get_sheet_by_name('TradesWithAdditionalInfo')

    trade_id_transaction_dict = get_trade_id_transaction_dict(trades_sheet)
    dict_of_commissions = get_trade_id_to_commission_dict(work_book)
    output_lines = add_commission(dict_of_commissions, trade_id_transaction_dict)
    print_output_lines(output_lines)


def print_output_lines(output_lines):
    headers = ["Unique Identifier", "Trade Date", "Instrument Code", "Market Code", "Quantity", "Price", "Transaction Type",
               "Exchange Rate", "Brokerage", "Brokerage Currency", "Comments"]
    print(','.join(headers))
    for t in output_lines:
        print(t.to_csv())

def main_ig(csv_file):
    output_lines = []
    with open(csv_file, 'rb') as csvfile:
        file_reader = csv.reader(csvfile, delimiter=',', quotechar='|')
        for row in file_reader:
            if row[1] != "Time" :
                transaction = create_transaction_for_ig(row)
                output_lines.append(transaction)
    output_lines.sort(key=lambda x: x.trade_id)
    print_output_lines(output_lines)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Transforming SaxoTrades to ShareSight format')
    parser.add_argument('--file', metavar='path', required=True,
                        help='the Excel sheet export of trades from Saxo Markets. In SaxoTrader, you can find this in Account > Historical reports > Trades. ' +
                             'Click "Open", the funnel icon on the left, then export.')
    parser.add_argument('--broker', metavar='path', required=True, default='saxo', choices=['saxo', 'ig'], help='Valid options "saxo" for SaxoBank or "ig" for IG. Default, saxo')
    args = parser.parse_args()
    if args.broker == "saxo":
        main_saxo(args.file)
    elif args.broker == "ig" :
        main_ig(args.file)
    else:
        raise Exception("Unsupported broker: " + args.broker)
